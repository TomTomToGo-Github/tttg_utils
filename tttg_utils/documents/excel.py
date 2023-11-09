## built-in modules
import string
import re
import os
from pathlib import Path
## pip installed modules
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
## selfmade modules


class ExcelPowerhouse:
    def __init__(self, path_file=None, sheets=None):
        self.path_file = Path(path_file or os.getcwd())
        self.sheets = sheets

    @staticmethod
    def new_xlsx_file(path_file: str):
        path_file = Path(path_file)
        if not path_file.suffix:
            path_file = Path(str(path_file) + ".xlsx")
        elif path_file.suffix != ".xlsx":
            print(f"Incorrect file extension: {path_file.suffix}. File not created")
            # quit()
        if not path_file.parent.is_dir():
            path_file.parent.mkdir(parents=True)
        if not path_file.is_file():
            workbook = openpyxl.Workbook(path_file)
            workbook.save(path_file)
            workbook.close()

    @staticmethod
    def new_xlsx_sheet(path_file, sheet_name=None):
        workbook = openpyxl.load_workbook(path_file)
        if not sheet_name:
            default_sheet_name = "Sheet"
            if default_sheet_name in workbook.sheetnames:
                sheet_postfixes = [re.sub(default_sheet_name, "", shn) for shn in workbook.sheetnames]
                new_sheet_num = max([int(sh or "0") for sh in sheet_postfixes if re.match("^(?=$)|\d+", sh)]) + 1
                sheet_name = default_sheet_name + str(new_sheet_num)
        # Load the Excel file
        # Create a new sheet in the same workbook and copy contents
        if sheet_name not in workbook.sheetnames:
            workbook.create_sheet(sheet_name)
        else:
            print(f"Sheet {sheet_name} already exists -> leaving it untouched")
        # Save the modified workbook
        workbook.save(path_file)
        workbook.close()

    @staticmethod
    def new_xlsx_sheet_from_template(path_file, template_sheet_name=None, new_sheet_name=None):
        # Load the Excel file
        workbook = openpyxl.load_workbook(path_file)
        template_sheet = workbook[template_sheet_name]
        # Create a new sheet in the same workbook and copy contents
        if new_sheet_name not in workbook.sheetnames:
            destination_sheet = workbook.copy_worksheet(template_sheet)
            destination_sheet.title = new_sheet_name
        # Save the modified workbook
        workbook.save(path_file)
        workbook.close()

    @staticmethod
    def compute_bool_filters(concat_df: pd.DataFrame, filters=None):
        bool_input_filters = pd.Series([True] * len(concat_df))
        if filters:
            nonempty_filters = {key: value for key, value in filters.items() if value}
            if nonempty_filters:  # handles case where user input all None...
                for col, filters_list in nonempty_filters.items():
                    if col not in concat_df:
                        print(f"\n\nColumn to filter on does not exist. Choose from {list(concat_df)}")
                    new_col_filter = pd.Series([] * len(concat_df))
                    for filter_value, filter_type, filter_agg in filters_list:
                        ## Filter by type
                        if filter_type.strip("s") == "contain":
                            new_filter = concat_df[col].str.contains(filter_value)
                        elif filter_type.strip("s") == "equal":
                            new_filter = concat_df[col].eq(filter_value)
                        else:
                            print(f"Specified filter ({filter_type}) does not exist (Options: equals, contains)")
                        ## Aggregate filters
                        if filter_agg == "and":
                            if new_col_filter.empty:
                                new_col_filter = new_filter
                            else:
                                new_col_filter = new_col_filter & new_filter
                        if filter_agg == "or":
                            if new_col_filter.empty:
                                new_col_filter = new_filter
                            else:
                                new_col_filter = new_col_filter | new_filter
                    bool_input_filters = bool_input_filters & new_col_filter
        return bool_input_filters

    @staticmethod
    def write_data_to_excel_sheet(excel_data_dict, path_file, new_sheet_name=None):
        workbook = openpyxl.load_workbook(path_file)
        worksheet = workbook[new_sheet_name]

        # Color Problematic entries
        for row_number in excel_data_dict.pop("problem_rows"):
            fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")  # light red
            # Apply the fill color to the entire row
            for cell in worksheet[row_number]:
                cell.fill = fill

        for cell, data in excel_data_dict.items():
            if isinstance(data, str):
                worksheet[cell] = data
                # worksheet['G8'] = 70
            elif isinstance(data, pd.DataFrame):
                ## Close previous writer
                workbook.save(path_file)
                workbook.close()
                ## Identify the starting point of the DataFrame
                potential_cols = [char for char in string.ascii_uppercase] + [
                    char1 + char2 for char1 in string.ascii_uppercase for char2 in string.ascii_uppercase
                ]
                startrow = int(re.sub("[a-zA-Z]", "", cell))
                startcol = potential_cols.index(re.sub("\d", "", cell))
                ## Write excel_monthly_df to file
                writer = pd.ExcelWriter(path_file, engine="openpyxl", mode="a", if_sheet_exists="overlay")
                data.to_excel(writer, sheet_name=new_sheet_name, index=False, startrow=startrow, startcol=startcol)
                writer._save()
                writer.close()
                ## Reopen previous writer
                workbook = openpyxl.load_workbook(path_file)
                worksheet = workbook[new_sheet_name]
        workbook.save(path_file)
        workbook.close()
    
    def ensure_xlsx_sheet_structure(self, file_xlsx=None, sheets=None):
        if not file_xlsx.is_file():
            self.new_xlsx_file(file_xlsx)
        if not sheets and self.sheets:
            sheets = self.sheets
        for sheet in sheets:
            self.new_xlsx_sheet(file_xlsx, sheet_name=sheet)
